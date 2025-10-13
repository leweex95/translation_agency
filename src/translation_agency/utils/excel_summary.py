"""Excel summary generator for translation pipeline validation tracking."""

import re
from pathlib import Path
from typing import Dict, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


class ExcelSummaryGenerator:
    """Generate Excel summary showing sentence-level changes across pipeline steps."""

    # Green fill for cells that changed
    CHANGE_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    @staticmethod
    def split_into_sentences(text: str) -> List[str]:
        """
        Split text into sentences using regex.
        Handles common sentence endings and preserves formatting.
        """
        # Split on sentence endings but keep the punctuation
        sentences = re.split(r'(?<=[.!?])\s+', text.strip())

        # Filter out empty sentences and clean up
        sentences = [s.strip() for s in sentences if s.strip()]

        return sentences

    @classmethod
    def generate_summary(cls, output_dir: str, document_name: str) -> str:
        """
        Generate Excel summary file showing sentence progression across all steps.

        Args:
            output_dir: Base output directory
            document_name: Name of the document being processed

        Returns:
            Path to the generated Excel file
        """
        document_dir = Path(output_dir) / document_name

        # Read all step files
        step_files = sorted(document_dir.glob("step*_*.txt"))
        final_file = document_dir / "final.txt"

        if not step_files:
            raise ValueError(f"No step files found in {document_dir}")

        # Read content from each step
        step_contents = {}
        for step_file in step_files:
            step_name = step_file.stem  # e.g., "step1_initial_translation", "step2_grammar_validation", etc.
            try:
                content = step_file.read_text(encoding='utf-8')
                step_contents[step_name] = content
            except Exception as e:
                print(f"Warning: Could not read {step_file}: {e}")
                continue

        # Also include final if it exists and is different
        if final_file.exists():
            try:
                final_content = final_file.read_text(encoding='utf-8')
                # Only add if different from last step
                last_step = f"step{len(step_contents)}_crossllm_validation"
                if last_step not in step_contents or step_contents[last_step] != final_content:
                    step_contents["final"] = final_content
            except Exception as e:
                print(f"Warning: Could not read final.txt: {e}")

        if not step_contents:
            raise ValueError("No valid step content found")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sentence Progression"

        # Get all step names in order
        def sort_key(step_name):
            if step_name == 'final':
                return 999
            # Extract step number from "step1_initial_translation" -> 1
            import re
            match = re.match(r'step(\d+)', step_name)
            return int(match.group(1)) if match else 999
        
        step_names = sorted(step_contents.keys(), key=sort_key)

        # Write headers - make them more readable
        ws['A1'] = "Sentence #"
        for col_num, step_name in enumerate(step_names, 2):
            if step_name == 'final':
                display_name = "Final"
            else:
                # Convert "step1_initial_translation" to "Step 1: Initial Translation"
                parts = step_name.split('_', 2)
                if len(parts) >= 3:
                    step_num = parts[0].replace('step', '')
                    step_type = parts[1]
                    step_desc = parts[2].replace('_', ' ')
                    display_name = f"Step {step_num}: {step_type.title()} {step_desc.title()}"
                else:
                    display_name = step_name.replace('_', ' ').title()
            ws.cell(row=1, column=col_num, value=display_name)

        # Process sentences for each step
        max_sentences = 0
        step_sentences = {}

        for step_name, content in step_contents.items():
            sentences = cls.split_into_sentences(content)
            step_sentences[step_name] = sentences
            max_sentences = max(max_sentences, len(sentences))

        # Write sentence data and track changes
        for sentence_idx in range(max_sentences):
            # Write sentence number
            ws.cell(row=sentence_idx + 2, column=1, value=sentence_idx + 1)

            prev_sentence = None
            for col_num, step_name in enumerate(step_names, 2):
                sentences = step_sentences[step_name]
                current_sentence = sentences[sentence_idx] if sentence_idx < len(sentences) else ""

                # Write the sentence
                cell = ws.cell(row=sentence_idx + 2, column=col_num, value=current_sentence)

                # Highlight if changed from previous step
                if prev_sentence is not None and current_sentence != prev_sentence:
                    cell.fill = cls.CHANGE_FILL

                prev_sentence = current_sentence

        # Auto-adjust column widths
        for col_num in range(1, len(step_names) + 2):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row_num in range(1, max_sentences + 2):
                cell_value = str(ws.cell(row=row_num, column=col_num).value or "")
                max_length = max(max_length, len(cell_value))

            # Set column width (min 10, max 50)
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)

        # Save the Excel file
        excel_path = document_dir / "pipeline_summary.xlsx"
        wb.save(excel_path)

        return str(excel_path)

    @classmethod
    def has_changes(cls, step_contents: Dict[str, str]) -> bool:
        """
        Check if there are any changes between steps.

        Args:
            step_contents: Dictionary mapping step names to content

        Returns:
            True if any changes detected between consecutive steps
        """
        def sort_key(step_name):
            if step_name == 'final':
                return 999
            # Extract step number from "step1_initial_translation" -> 1
            import re
            match = re.match(r'step(\d+)', step_name)
            return int(match.group(1)) if match else 999
        
        step_names = sorted(step_contents.keys(), key=sort_key)

        for i in range(1, len(step_names)):
            prev_content = step_contents[step_names[i-1]]
            curr_content = step_contents[step_names[i]]

            if prev_content != curr_content:
                return True

        return False